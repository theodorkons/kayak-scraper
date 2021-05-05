from time import sleep
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from datetime import timedelta
from openpyxl import load_workbook 
import openpyxl

sleep_time = 10

city_from = 'OSL'
city_to = 'ATH'
date = datetime(2021, 9, 1)
string_date = date
string_date = date.strftime('%d-%m-%Y')
until = datetime(2021, 9, 30)
until = until.strftime('%d-%m-%Y')

#excel_name = 'fromAthens'+datetime.today().strftime('%d-%m-%Y')
excel_name = 'toAthens'+datetime.today().strftime('%d-%m-%Y')

delta_date = datetime.strptime(string_date, '%d-%m-%Y')
delta_until = datetime.strptime(until, '%d-%m-%Y')
delta = delta_until - delta_date  #calculate the day difference

cols = (['Date', 'Departure Airport', 'Departure Time', 'Arrival Airport', 'Arrival Time', 'Stops', 'Travel Duration', 'Price', 'Links'])
df = pd.DataFrame(columns=cols)
flights_df = pd.DataFrame(columns=cols)

options = webdriver.ChromeOptions() 
options.add_experimental_option("excludeSwitches", ["enable-logging"])
PATH = "C:\Program Files (x86)\chromedriver.exe"
driver = webdriver.Chrome(options=options, executable_path=PATH)

kayak = 'https://www.gr.kayak.com/flights/' + city_from + '-' + city_to + '/' + date.strftime('%Y-%m-%d') + '?sort=bestflight_a'


def scrape():
    try:

        body = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CLASS_NAME, "listBody "))
        )
        
        results = WebDriverWait(driver, 30).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, "Flights-Results-FlightPriceSection"))
        )

        flight_times = body.find_elements_by_class_name('time-pair')

        """ for i in flight_times:
            print(i.text) """

        departure_time_list =[]
        for i in range(0, len(flight_times), 2):
            if flight_times[i].text != '': 
                departure_time_list.append(flight_times[i].text)
                #print(flight_times[i].text)
        #print(f"departure_time_list  {len(departure_time_list)}")
        arrival_time_list =[]
        for i in range(1, len(flight_times), 2):
            if flight_times[i].text != '':
                arrival_time_list.append(flight_times[i].text)
                #print(flight_times[i].text)
        #print(f"arrival_time_list {len(arrival_time_list)}")
        
    
        airports = body.find_elements_by_class_name('airport-name')
        airports_list = []
        for i in range(0, len(airports)):
            if airports[i].text != '':
                airports_list.append(airports[i].text)
                #print(airports[i].text)
        #airports_list = [airport.text for airport in airports]         
        departure_airports_list = airports_list[::2]  
        #print(f"departure_airports_list {len(departure_airports_list)}")                #This separates the departure and arrival airports
        arrival_airports_list = airports_list[1::2]
        #print(f"arrival_airports_list {len(arrival_airports_list)}")
        
        stops = body.find_elements_by_class_name('stops-text')

        stops_list = []
        for i in range(0, len(stops)):
            if stops[i].text != '':
                stops_list.append(stops[i].text)
                #print(stops[i].text)
        #stops_list = [stop.text for stop in stops]
        #print(f"stops_list {len(stops_list)}")
        
        durations = body.find_elements_by_class_name('duration')

        durations_list = []
        for i in range(0, len(durations)):
            if durations[i].text !='':
                durations_list.append(durations[i].text)
                #print(durations[i].text)
        #durations =[duration.text for duration in durations]
        #print(f"durations_list {len(durations_list)}")

        prices_list = []
        for result in results:
            if result.text != '':
                prices = result.find_element_by_class_name('price-text')
                price = prices.text
                #print(price)
                price = price.replace('€', '')        #removes the '.' so they can be sorted as int
                if '.' in price:
                    price = price.replace('.', '')
                prices_list.append(int(price))
        #print(f"prices_list {len(prices_list)}")
        
        global string_date, date
        date_list = [string_date] * len(prices_list)
        date += timedelta(days=1)
        string_date = date.strftime('%d-%m-%Y')
        
        link = driver.current_url
        links_list = [link] * len(prices_list)
        #print(f"links_list {len(links_list)}")
        
        flights_df = pd.DataFrame({
            'Date': date_list,
            'Departure Airport': departure_airports_list,
            'Departure Time': departure_time_list,
            'Arrival Airport': arrival_airports_list,
            'Arrival Time': arrival_time_list,
            'Stops': stops_list,
            'Travel Duration': durations_list,
            'Price': prices_list,
            'Links': links_list
        })[cols]
        print('ops')

        return flights_df
        
    except:
        
        print('Something went wrong')
        #flights_df = scrape()
        #driver.quit()
        sleep(3)
        return pd.DataFrame()

def nextDay():
    next_day_button = '//button[contains(@id, "dateRangeInput") and contains (@class,"_irG ")]'
    driver.find_elements_by_xpath(next_day_button)[1].click()

    search_button = '//div[contains(@id, "submit") and contains (@class,"Common-Widgets-Button-StyleJamButton")]'
    driver.find_elements_by_xpath(search_button)[0].click()



driver.implicitly_wait(10)
driver.get(kayak)

xp_popup_close = '//button[contains(@id, "accept") and contains (@class,"Common-Widgets-Button-StyleJamButton")]'
driver.find_element_by_xpath(xp_popup_close).click()

for i in range(0, delta.days):
    sleep(sleep_time)
    flights_df = scrape()
    while flights_df.empty:
        flights_df = scrape()
    df = df.append(flights_df, ignore_index=True)
    nextDay()
    

df.sort_values(by=['Price'], inplace=True)
""" with pd.ExcelWriter(excel_name + '.xlsx', index = False,
                    mode='a') as writer:  
    df.to_excel(writer) """
df.to_excel(excel_name + '.xlsx', index=False)







#workbook = openpyxl.load_workbook(excel_name + '.xlsx')
#with pd.ExcelWriter(excel_name + '.xlsx') as writer: #pylint: disable=abstract-class-instantiated
""" print('ep')
    writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
    df.to_excel(writer) """
""" writer = pd.ExcelWriter(excel_name + '.xlsx', engine='openpyxl')
# try to open an existing workbook
writer.book = load_workbook(excel_name + '.xlsx')
# copy existing sheets
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
# read existing file
reader = pd.read_excel(excel_name + '.xlsx')
# write out the new sheet
df.to_excel(writer,index=False,header=False,startrow=len(reader)+1)
writer.close() """


driver.quit()





